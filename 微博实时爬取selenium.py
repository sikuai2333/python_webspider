'''
Descripttion: 
version: 
Author: sikuai
Date: 2022-07-05 12:07:29
LastEditors: sikuai
LastEditTime: 2022-09-19 22:05:39

'''
from lib2to3.pgen2 import driver
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from lxml import etree
import openpyxl

i = 4

search_url = 'https://m.weibo.cn/search?containerid=231522type%3D1%26t%3D10%26q%3D%23%E5%AE%89%E5%BE%BD%E7%9C%81%E5%A4%A7%E5%AD%A6%E7%94%9F%E5%9B%BD%E5%BA%86%E5%81%87%E6%9C%9F%E5%8F%AA%E6%9C%893%E5%A4%A9%23&extparam=%23%E5%AE%89%E5%BE%BD%E7%9C%81%E5%A4%A7%E5%AD%A6%E7%94%9F%E5%9B%BD%E5%BA%86%E5%81%87%E6%9C%9F%E5%8F%AA%E6%9C%893%E5%A4%A9%23&luicode=10000011&lfid=231522type%3D61%26q%3D%23%E5%AE%89%E5%BE%BD%E7%9C%81%E5%A4%A7%E5%AD%A6%E7%94%9F%E5%9B%BD%E5%BA%86%E5%81%87%E6%9C%9F%E5%8F%AA%E6%9C%893%E5%A4%A9%23%26t%3D10'

#去除selenium自动提示
options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-automation'])
#去除selenium检测
options.add_argument('--disable-blink-features=AutomationControlled')
options.add_argument('--disable-blink-features=AutomationControlledByDomain')
options.add_argument('--disable-blink-features=AutomationControlledByUrl')
options.add_argument('--disable-blink-features=AutomationControlledByUrlBlockList')
#去除证书
options.add_argument('-ignore-certificate-errors')
options.add_argument('-ignore -ssl-errors')
driver = webdriver.Chrome(options=options)
prefs = {"profile.default_content_setting_values.notifications": 2}
options.add_experimental_option("prefs", prefs)

driver.get(search_url)
time.sleep(3)
shishi = driver.find_element(By.XPATH,r'/html/body/div/div[1]/div[1]/div[3]/div[2]/div[1]/div/div/div/ul/li[2]/span')
shishi.click()
time.sleep(3)
workbook = openpyxl.load_workbook("guoqing.xlsx")
sheet = workbook["Sheet1"]
while True:
    i = i + 1
    print(i)
#    time.sleep(3)
    try:
        driver.execute_script('window.scrollTo(0,document.body.scrollHeight)')
        # 页面循环的我有点懵，整个也是靠试出来的，不知道有没有更好的方法
        # /html/body/div/div[1]/div[1]/div[4]/div/div/div/header/div/div/a/h3
        name = driver.find_element(By.XPATH,r'/html/body/div/div[1]/div[1]/div['+str(i)+']/div/div/div/header/div/div/a/h3').text
        msg = driver.find_element(By.XPATH,r'/html/body/div/div[1]/div[1]/div['+str(i)+']/div/div/div/article/div[2]/div[1]').text
        msg_time = driver.find_element(By.XPATH,r'/html/body/div/div[1]/div[1]/div['+str(i)+']/div/div/div/header/div/div/h4/span[1]').text
        device = driver.find_element(By.XPATH,r'/html/body/div/div[1]/div[1]/div['+str(i)+']/div/div/div/header/div/div/h4/span[2]').text
        msg = msg.replace("#安徽省大学生国庆假期只有3天#","")
        device = device.replace("来自","")
        device = device.replace("客户端","")
        print(name+"在"+msg_time+"用"+device+"说："+msg)
        sheet.cell(row=i,column=1,value=name)
        sheet.cell(row=i,column=2,value=msg_time)    
        sheet.cell(row=i,column=3,value=device)
        sheet.cell(row=i,column=4,value=msg)
        workbook.save("guoqing.xlsx")
    except:
        continue