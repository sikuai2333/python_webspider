'''
Descripttion: 
version: 
Author: sikuai
Date: 2022-09-20 06:49:09
LastEditors: sikuai
LastEditTime: 2022-09-20 22:46:47
'''
from cgitb import text
import requests
import json
import time
import re
import openpyxl
a = 0
page = 2
workbook = openpyxl.load_workbook("国庆微博.xlsx")
sheet = workbook["Sheet1"]
for page in range(2,10000):
    print(str(page)+"页")
    url = 'https://m.weibo.cn/api/container/getIndex?containerid=231522type%3D61%26q%3D%23%E5%AE%89%E5%BE%BD%E7%9C%81%E5%A4%A7%E5%AD%A6%E7%94%9F%E5%9B%BD%E5%BA%86%E5%81%87%E6%9C%9F%E5%8F%AA%E6%9C%893%E5%A4%A9%23%26t%3D10&extparam=%23%E5%AE%89%E5%BE%BD%E7%9C%81%E5%A4%A7%E5%AD%A6%E7%94%9F%E5%9B%BD%E5%BA%86%E5%81%87%E6%9C%9F%E5%8F%AA%E6%9C%893%E5%A4%A9%23&luicode=10000011&lfid=231522type%3D61%26q%3D%23%E5%AE%89%E5%BE%BD%E7%9C%81%E5%A4%A7%E5%AD%A6%E7%94%9F%E5%9B%BD%E5%BA%86%E5%81%87%E6%9C%9F%E5%8F%AA%E6%9C%893%E5%A4%A9%23%26t%3D10&page_type=searchall&page='+str(page)
    res = requests.get(url)
    res = str(res.text)
    # print(res.encode('utf-8', 'replace').decode('utf-8'))
    res = res.encode('utf-8', 'replace').decode('utf-8')
    data = json.loads(res, strict=False)
    cards = data['data']['cards']
    # time.sleep(1)
    try:
        for i in range(1,10):
            # print(i)
            print(a)
            mblog = cards[i]['mblog']
            mblog_text = mblog['text']
            user_id = mblog['id']
            name = mblog['user']['screen_name']
            region_name = mblog['region_name']
            device = mblog['source']
            # print(mblog_text)
            chinese = re.findall('[\u4e00-\uFF1F]', mblog_text)
            chinese = str(chinese)
            chinese = chinese.replace(",","")
            chinese = chinese.replace("'","")
            chinese = chinese.replace("[","")
            chinese = chinese.replace("]","")
            chinese = chinese.replace(" ","")
            chinese = chinese.replace("安徽省大学生国庆假期只有天","")
            if chinese != '':
                a = a + 1
                print(device+region_name+user_id+name+chinese)
                # nws.write(0,2,"dd")
                # nws.write(a,0,device)
                # nws.write(a,1,region_name)
                # nws.write(a,2,user_id)
                # nws.write(a,3,name)
                # nws.write(a,4,chinese)
                sheet.cell(row=a,column=1,value=device)
                sheet.cell(row=a,column=2,value=region_name)
                sheet.cell(row=a,column=3,value=user_id)
                sheet.cell(row=a,column=4,value=name)
                sheet.cell(row=a,column=5,value=chinese)
                workbook.save("国庆微博.xlsx")

    except:
        continue



